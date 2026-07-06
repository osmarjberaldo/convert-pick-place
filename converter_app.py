#!/usr/bin/env python3
"""
PCB Pick and Place Converter
Converts Altium Pick and Place files to:
  - POCISAO format (XLSX/XML)
  - BOM format (XLSX) grouped by component value
Cross-platform GUI (Windows & Linux).
"""

import os
import sys
import csv
import xml.etree.ElementTree as ET
from xml.dom import minidom
from tkinter import (
    Tk, Toplevel, Frame, Label, Button, Text, Canvas, Scrollbar, filedialog,
    messagebox, ttk, StringVar, BooleanVar, Checkbutton, Radiobutton,
    Menu
)
from tkinter import TclError
import subprocess
import webbrowser
from urllib.parse import quote

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# =============================================================================
# Translation System (PT-BR primary, EN secondary)
# =============================================================================

LANGUAGES = {
    "pt_BR": {
        "app_title": "PCB Pick and Place Converter | by DevOsmar",
        "app_subtitle": "by DevOsmar",
        "menu_file": "Arquivo",
        "menu_exit": "Sair",
        "menu_help": "Ajuda",
        "menu_about": "Sobre",
        "menu_language": "Idioma",
        "lang_pt": "Português (BR)",
        "lang_en": "English",
        "tab_pocisao": "  \U0001f4cd Pick Place \u2192 POCISAO  ",
        "tab_bom": "  \U0001f4cb Pick Place \u2192 BOM  ",
        "pocisao_header": "Pick Place \u2192 PC POCISAO",
        "pocisao_desc": "Converte coordenadas X/Y para formato de montagem",
        "pocisao_desc_en": "Converts X/Y coordinates to assembly format",
        "input_file": "Arquivo de Entrada",
        "input_file_label": "Arquivo:",
        "btn_select": "Selecionar",
        "btn_select_en": "Browse",
        "preview_data": "Preview dos Dados",
        "preview_data_en": "Data Preview",
        "no_file_loaded": "Nenhum arquivo carregado",
        "no_file_loaded_en": "No file loaded",
        "btn_add_row": "\u2795 Adicionar Linha",
        "btn_add_row_en": "\u2795 Add Row",
        "btn_delete_row": "\U0001f5d1\ufe0f Excluir Linha",
        "btn_delete_row_en": "\U0001f5d1\ufe0f Delete Row",
        "output_options": "Op\u00e7\u00f5es de Sa\u00edda",
        "output_options_en": "Output Options",
        "format_label": "Formato:",
        "format_xlsx": "XLSX (Excel)",
        "format_xml": "XML",
        "include_metadata": "Incluir metadados (XML)",
        "include_metadata_en": "Include metadata (XML)",
        "btn_convert": "Converter e Salvar...",
        "btn_convert_en": "Convert and Save...",
        "dialog_select_pnp": "Selecione o arquivo Pick and Place",
        "dialog_select_pnp_en": "Select Pick and Place file",
        "supported_files": "Arquivos suportados",
        "supported_files_en": "Supported files",
        "all_files": "Todos",
        "all_files_en": "All files",
        "units": "Unidades",
        "units_en": "Units",
        "components": "componente(s)",
        "components_en": "component(s)",
        "no_data": "Nenhum dado",
        "no_data_en": "No data",
        "menu_edit": "\u270f\ufe0f Editar",
        "menu_edit_en": "\u270f\ufe0f Edit",
        "warn_no_data": "Nenhum dado",
        "warn_no_data_en": "No data",
        "warn_load_first": "Carregue um arquivo primeiro.",
        "warn_load_first_en": "Load a file first.",
        "dialog_save_as": "Salvar como",
        "dialog_save_as_en": "Save as",
        "file_saved": "Arquivo salvo com sucesso!",
        "file_saved_en": "File saved successfully!",
        "dialog_complete": "Conclu\u00eddo",
        "dialog_complete_en": "Completed",
        "ask_open_file": "Abrir arquivo?",
        "ask_open_file_en": "Open file?",
        "error_title": "Erro",
        "error_title_en": "Error",
        "error_convert": "Erro ao converter:",
        "error_convert_en": "Conversion error:",
        "error_load": "Erro ao carregar",
        "error_load_en": "Error loading",
        "unsupported_format": "Formato n\u00e3o suportado",
        "unsupported_format_en": "Unsupported format",
        # BOM tab
        "bom_header": "Pick Place \u2192 BOM (Lista de Componentes)",
        "bom_header_en": "Pick Place \u2192 BOM (Component List)",
        "bom_desc": "Agrupa componentes iguais para lista de compras JLCPCB",
        "bom_desc_en": "Groups identical components for JLCPCB shopping list",
        "input_file_csv": "Arquivo de Entrada (CSV)",
        "input_file_csv_en": "Input File (CSV)",
        "preview_bom": "Preview BOM (Agrupado por valor)",
        "preview_bom_en": "BOM Preview (Grouped by value)",
        "component_groups": "grupo(s) de componentes",
        "component_groups_en": "component group(s)",
        "dialog_save_bom": "Salvar BOM como",
        "dialog_save_bom_en": "Save BOM as",
        "bom_saved": "BOM salva com sucesso!",
        "bom_saved_en": "BOM saved successfully!",
        # About dialog
        "about_title": "Sobre - PCB Pick and Place Converter",
        "about_title_en": "About - PCB Pick and Place Converter",
        "about_version": "v1.0 \u00b7 Open Source",
        "about_desc": "Converte arquivos Altium Pick and Place\npara formato PC POCISAO (XLSX/XML) e BOM.\nFerramenta para automa\u00e7\u00e3o de montagem de PCB.",
        "about_desc_en": "Converts Altium Pick and Place files\nto PC POCISAO format (XLSX/XML) and BOM.\nPCB assembly automation tool.",
        "about_created_by": "Criado por",
        "about_created_by_en": "Created by",
        "about_oss": "Este software \u00e9 Open Source - contribui\u00e7\u00f5es s\u00e3o bem-vindas!",
        "about_oss_en": "This software is Open Source - contributions are welcome!",
        "about_donate": "\u2615 Gostou do projeto? Contribua!",
        "about_donate_en": "\u2615 Enjoy the project? Contribute!",
        "btn_paypal": "\U0001f4b8 Doar com PayPal",
        "btn_paypal_en": "\U0001f4b8 Donate with PayPal",
        "btn_pix": "\U0001f1e7\U0001f1f7 Doar com PIX",
        "btn_pix_en": "\U0001f1e7\U0001f1f7 Donate with PIX",
        "pix_label": "Chave PIX: ",
        "pix_label_en": "PIX Key: ",
        "btn_close": "Fechar",
        "btn_close_en": "Close",
        "pix_copied_title": "Chave PIX copiada!",
        "pix_copied_title_en": "PIX Key copied!",
        "pix_copied_msg": "Chave PIX copiada para a \u00e1rea de transfer\u00eancia:\n\n{key}\n\nCole no seu aplicativo do banco para fazer a doa\u00e7\u00e3o.",
        "pix_copied_msg_en": "PIX Key copied to clipboard:\n\n{key}\n\nPaste into your banking app to donate.",
        # Dependencies
        "dep_title": "Depend\u00eancia Necess\u00e1ria",
        "dep_title_en": "Required Dependency",
        "dep_msg": "A biblioteca 'openpyxl' \u00e9 necess\u00e1ria para exportar arquivos XLSX (Excel).\n\nDeseja instalar automaticamente agora?",
        "dep_msg_en": "The 'openpyxl' library is required to export XLSX (Excel) files.\n\nDo you want to install it now?",
        "dep_install_error": "Erro na Instala\u00e7\u00e3o",
        "dep_install_error_en": "Installation Error",
        "dep_install_msg": "N\u00e3o foi poss\u00edvel instalar o openpyxl automaticamente.\n\nExecute manualmente no terminal:\npip install openpyxl",
        "dep_install_msg_en": "Could not install openpyxl automatically.\n\nRun manually in terminal:\npip install openpyxl",
        # File info / status
        "file_info_txt": "{file} | {count} {components} | {units_label}: {units}",
        "file_info_txt_en": "{file} | {count} {components} | {units_label}: {units}",
        "file_info_bom": "{file} | {raw} componentes \u2192 {count} grupos",
        "file_info_bom_en": "{file} | {raw} components \u2192 {count} groups",
        "file_saved_msg": "Arquivo salvo com sucesso!\n\n{path}\n\n{count} {components}.",
        "file_saved_msg_en": "File saved successfully!\n\n{path}\n\n{count} {components}.",
        "bom_saved_msg": "BOM salva com sucesso!\n\n{path}\n\n{count} {groups}.",
        "bom_saved_msg_en": "BOM saved successfully!\n\n{path}\n\n{count} {groups}.",
        "select_pnp_title": "Selecione o arquivo Pick and Place",
        "select_pnp_title_en": "Select Pick and Place file",
        "select_csv_title": "Selecione o arquivo CSV do Altium",
        "select_csv_title_en": "Select Altium CSV file",
        "btn_search": "🔍 Buscar Componente",
        "search_octopart": "Octopart",
        "search_snapmagic": "SnapMagic",
        "search_mouser": "Mouser (BR)",
        "search_jlcpcb": "JLCPCB Parts",
        "search_digikey": "DigiKey",
        "search_snapeda": "SnapEDA",
        "search_lcsc": "LCSC",
        "search_no_selection": "Selecione um componente na tabela primeiro.",
        "menu_search": "🔍 Buscar em",
        # Tooltips
        "status_ready": "Pronto",
        "status_file_loaded": "Arquivo carregado: {file}",
        "status_saved": "Salvo: {file}",
        "status_error": "Erro: {msg}",
        "tip_select_file": "Selecionar arquivo Pick and Place",
        "tip_add_row": "Adicionar nova linha",
        "tip_delete_row": "Excluir linha selecionada",
        "tip_convert": "Converter e salvar arquivo",
        "tip_search_octopart": "Buscar componente no Octopart",
        "tip_search_snapmagic": "Buscar componente no SnapMagic",
        "tip_search_mouser": "Buscar componente na Mouser",
        "tip_search_jlcpcb": "Buscar componente no JLCPCB Parts",
        "tip_search_digikey": "Buscar componente na DigiKey",
        "tip_search_snapeda": "Buscar footprint/símbolo no SnapEDA",
        "tip_search_lcsc": "Buscar componente no LCSC",
        "tip_paypal": "Doar via PayPal (internacional)",
        "tip_pix": "Doar via PIX (Brasil)",
        "tip_dev_profile": "Ver perfil do DevOsmar no GitHub",
        "tip_close": "Fechar janela",
        "open_file_error": "Abrir arquivo",
        "open_file_error_en": "Open file",
    },
    "en_US": {
        "app_title": "PCB Pick and Place Converter | by DevOsmar",
        "app_subtitle": "by DevOsmar",
        "menu_file": "File",
        "menu_exit": "Exit",
        "menu_help": "Help",
        "menu_about": "About",
        "menu_language": "Language",
        "lang_pt": "Portugu\u00eas (BR)",
        "lang_en": "English",
        "tab_pocisao": "  \U0001f4cd Pick Place \u2192 POCISAO  ",
        "tab_bom": "  \U0001f4cb Pick Place \u2192 BOM  ",
        "pocisao_header": "Pick Place \u2192 PC POCISAO",
        "pocisao_desc": "Converts X/Y coordinates to assembly format",
        "pocisao_desc_en": "Converts X/Y coordinates to assembly format",
        "input_file": "Input File",
        "input_file_label": "File:",
        "btn_select": "Browse",
        "btn_select_en": "Browse",
        "preview_data": "Data Preview",
        "preview_data_en": "Data Preview",
        "no_file_loaded": "No file loaded",
        "no_file_loaded_en": "No file loaded",
        "btn_add_row": "\u2795 Add Row",
        "btn_add_row_en": "\u2795 Add Row",
        "btn_delete_row": "\U0001f5d1\ufe0f Delete Row",
        "btn_delete_row_en": "\U0001f5d1\ufe0f Delete Row",
        "output_options": "Output Options",
        "output_options_en": "Output Options",
        "format_label": "Format:",
        "format_xlsx": "XLSX (Excel)",
        "format_xml": "XML",
        "include_metadata": "Include metadata (XML)",
        "include_metadata_en": "Include metadata (XML)",
        "btn_convert": "Convert and Save...",
        "btn_convert_en": "Convert and Save...",
        "dialog_select_pnp": "Select Pick and Place file",
        "dialog_select_pnp_en": "Select Pick and Place file",
        "supported_files": "Supported files",
        "supported_files_en": "Supported files",
        "all_files": "All files",
        "all_files_en": "All files",
        "units": "Units",
        "units_en": "Units",
        "components": "component(s)",
        "components_en": "component(s)",
        "no_data": "No data",
        "no_data_en": "No data",
        "menu_edit": "\u270f\ufe0f Edit",
        "menu_edit_en": "\u270f\ufe0f Edit",
        "warn_no_data": "No data",
        "warn_no_data_en": "No data",
        "warn_load_first": "Load a file first.",
        "warn_load_first_en": "Load a file first.",
        "dialog_save_as": "Save as",
        "dialog_save_as_en": "Save as",
        "file_saved": "File saved successfully!",
        "file_saved_en": "File saved successfully!",
        "dialog_complete": "Completed",
        "dialog_complete_en": "Completed",
        "ask_open_file": "Open file?",
        "ask_open_file_en": "Open file?",
        "error_title": "Error",
        "error_title_en": "Error",
        "error_convert": "Conversion error:",
        "error_convert_en": "Conversion error:",
        "error_load": "Error loading",
        "error_load_en": "Error loading",
        "unsupported_format": "Unsupported format",
        "unsupported_format_en": "Unsupported format",
        "bom_header": "Pick Place \u2192 BOM (Component List)",
        "bom_header_en": "Pick Place \u2192 BOM (Component List)",
        "bom_desc": "Groups identical components for JLCPCB shopping list",
        "bom_desc_en": "Groups identical components for JLCPCB shopping list",
        "input_file_csv": "Input File (CSV)",
        "input_file_csv_en": "Input File (CSV)",
        "preview_bom": "BOM Preview (Grouped by value)",
        "preview_bom_en": "BOM Preview (Grouped by value)",
        "component_groups": "component group(s)",
        "component_groups_en": "component group(s)",
        "dialog_save_bom": "Save BOM as",
        "dialog_save_bom_en": "Save BOM as",
        "bom_saved": "BOM saved successfully!",
        "bom_saved_en": "BOM saved successfully!",
        "about_title": "About - PCB Pick and Place Converter",
        "about_title_en": "About - PCB Pick and Place Converter",
        "about_version": "v1.0 \u00b7 Open Source",
        "about_desc": "Converts Altium Pick and Place files\nto PC POCISAO format (XLSX/XML) and BOM.\nPCB assembly automation tool.",
        "about_desc_en": "Converts Altium Pick and Place files\nto PC POCISAO format (XLSX/XML) and BOM.\nPCB assembly automation tool.",
        "about_created_by": "Created by",
        "about_created_by_en": "Created by",
        "about_oss": "This software is Open Source - contributions are welcome!",
        "about_oss_en": "This software is Open Source - contributions are welcome!",
        "about_donate": "\u2615 Enjoy the project? Contribute!",
        "about_donate_en": "\u2615 Enjoy the project? Contribute!",
        "btn_paypal": "\U0001f4b8 Donate with PayPal",
        "btn_paypal_en": "\U0001f4b8 Donate with PayPal",
        "btn_pix": "\U0001f4b8 Donate with PIX",
        "btn_pix_en": "\U0001f4b8 Donate with PIX",
        "pix_label": "PIX Key: ",
        "pix_label_en": "PIX Key: ",
        "btn_close": "Close",
        "btn_close_en": "Close",
        "pix_copied_title": "PIX Key copied!",
        "pix_copied_title_en": "PIX Key copied!",
        "pix_copied_msg": "PIX Key copied to clipboard:\n\n{key}\n\nPaste into your banking app to donate.",
        "pix_copied_msg_en": "PIX Key copied to clipboard:\n\n{key}\n\nPaste into your banking app to donate.",
        "dep_title": "Required Dependency",
        "dep_title_en": "Required Dependency",
        "dep_msg": "The 'openpyxl' library is required to export XLSX (Excel) files.\n\nDo you want to install it now?",
        "dep_msg_en": "The 'openpyxl' library is required to export XLSX (Excel) files.\n\nDo you want to install it now?",
        "dep_install_error": "Installation Error",
        "dep_install_error_en": "Installation Error",
        "dep_install_msg": "Could not install openpyxl automatically.\n\nRun manually in terminal:\npip install openpyxl",
        "dep_install_msg_en": "Could not install openpyxl automatically.\n\nRun manually in terminal:\npip install openpyxl",
        "file_info_txt": "{file} | {count} {components} | {units_label}: {units}",
        "file_info_txt_en": "{file} | {count} {components} | {units_label}: {units}",
        "file_info_bom": "{file} | {raw} components \u2192 {count} groups",
        "file_info_bom_en": "{file} | {raw} components \u2192 {count} groups",
        "file_saved_msg": "File saved successfully!\n\n{path}\n\n{count} {components}.",
        "file_saved_msg_en": "File saved successfully!\n\n{path}\n\n{count} {components}.",
        "bom_saved_msg": "BOM saved successfully!\n\n{path}\n\n{count} {groups}.",
        "bom_saved_msg_en": "BOM saved successfully!\n\n{path}\n\n{count} {groups}.",
        "select_pnp_title": "Select Pick and Place file",
        "select_pnp_title_en": "Select Pick and Place file",
        "select_csv_title": "Select Altium CSV file",
        "select_csv_title_en": "Select Altium CSV file",
        "btn_search": "🔍 Search Component",
        "btn_search_en": "🔍 Search Component",
        "search_octopart": "Octopart",
        "search_snapmagic": "SnapMagic",
        "search_mouser": "Mouser (BR)",
        "search_jlcpcb": "JLCPCB Parts",
        "search_digikey": "DigiKey",
        "search_snapeda": "SnapEDA",
        "search_lcsc": "LCSC",
        "search_no_selection": "Select a component in the table first.",
        "search_no_selection_en": "Select a component in the table first.",
        "menu_search": "🔍 Search on",
        "menu_search_en": "🔍 Search on",
        # Tooltips
        "status_ready": "Ready",
        "status_file_loaded": "File loaded: {file}",
        "status_saved": "Saved: {file}",
        "status_error": "Error: {msg}",
        "tip_select_file": "Browse Pick and Place file",
        "tip_add_row": "Add new row",
        "tip_delete_row": "Delete selected row",
        "tip_convert": "Convert and save file",
        "tip_search_octopart": "Search component on Octopart",
        "tip_search_snapmagic": "Search component on SnapMagic",
        "tip_search_mouser": "Search component on Mouser",
        "tip_search_jlcpcb": "Search component on JLCPCB Parts",
        "tip_search_digikey": "Search component on DigiKey",
        "tip_search_snapeda": "Search footprint/symbol on SnapEDA",
        "tip_search_lcsc": "Search component on LCSC",
        "tip_paypal": "Donate via PayPal (international)",
        "tip_pix": "Donate via PIX (Brazil)",
        "tip_dev_profile": "View DevOsmar's GitHub profile",
        "tip_close": "Close window",
        "open_file_error": "Open file",
        "open_file_error_en": "Open file",
    },
}


def detect_system_language():
    """Detect the system locale and return 'pt_BR' or 'en_US'."""
    # Try Windows API (Python 3.13+ where getdefaultlocale is removed)
    try:
        import ctypes
        lang_id = ctypes.windll.kernel32.GetUserDefaultUILanguage()
        # Primary language ID for Portuguese = 0x16 (mask: low 10 bits)
        if (lang_id & 0x3FF) == 0x16:
            return "pt_BR"
    except Exception:
        pass
    # Try locale module (cross-platform, deprecated in 3.11 but works on ≤3.12)
    try:
        import locale
        lang_code, _ = locale.getdefaultlocale()
        if lang_code and lang_code.startswith("pt"):
            return "pt_BR"
    except Exception:
        pass
    # Fallback: check environment variables (Linux/macOS)
    for var in ("LANG", "LC_ALL", "LC_MESSAGES"):
        val = os.environ.get(var, "")
        if val.lower().startswith("pt"):
            return "pt_BR"
    return "en_US"


class Lang:
    """Simple translation manager. Defaults to PT-BR, falls back to key."""

    def __init__(self, lang=None):
        if lang is None:
            lang = detect_system_language()
        self.lang = lang
        self._listeners = []

    def t(self, key, **kwargs):
        """Translate key. Supports {placeholder} formatting via kwargs."""
        lang_dict = LANGUAGES.get(self.lang, LANGUAGES["pt_BR"])
        val = lang_dict.get(key)
        if val is None:
            val = LANGUAGES["pt_BR"].get(key, key)
        if kwargs:
            try:
                val = val.format(**kwargs)
            except KeyError:
                pass
        return val

    def switch_to(self, lang):
        """Switch language and notify listeners."""
        if lang in LANGUAGES and lang != self.lang:
            self.lang = lang
            for cb in self._listeners:
                try:
                    cb()
                except Exception:
                    pass

    def add_listener(self, callback):
        """Register a callback to be called when language changes."""
        self._listeners.append(callback)


# =============================================================================
# Parsing Engine
# =============================================================================

# Column positions from the Altium Pick and Place TXT header:
# Designator(0-11) Comment(11-27) Layer(27-36) Footprint(36-61)
# Center-X(mm)(61-74) Center-Y(mm)(74-87) Rotation(87-96) Description(96+)


def detect_encoding(filepath):
    """Detect file encoding by trying common encodings."""
    with open(filepath, "rb") as f:
        raw = f.read(65536)
    for enc in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
        try:
            raw.decode(enc)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return "latin-1"


def parse_pick_place_txt(filepath):
    """
    Parse an Altium Pick and Place TXT file (fixed-width columns).
    Returns (metadata, rows).
    """
    encoding = detect_encoding(filepath)
    with open(filepath, "r", encoding=encoding) as f:
        content = f.read()

    lines = content.splitlines()
    metadata = {}
    header_line_idx = None

    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith("Date:"):
            metadata["Date"] = stripped.split("Date:")[-1].strip()
        elif stripped.startswith("Time:"):
            metadata["Time"] = stripped.split("Time:")[-1].strip()
        elif stripped.startswith("Revision:"):
            metadata["Revision"] = stripped.split("Revision:")[-1].strip()
        elif stripped.startswith("Variant:"):
            metadata["Variant"] = stripped.split("Variant:")[-1].strip()
        elif stripped.startswith("Units used:"):
            metadata["Units"] = stripped.split("Units used:")[-1].strip()

        if "Designator" in stripped and "Center-X" in stripped:
            header_line_idx = i
            break

    if header_line_idx is None:
        raise ValueError("Could not find Pick and Place header line in the file.")

    header_line = lines[header_line_idx]
    col_names = ["Designator", "Comment", "Layer", "Footprint",
                 "Center-X(mm)", "Center-Y(mm)", "Rotation", "Description"]
    col_positions = [header_line.find(c) for c in col_names]
    col_positions.append(len(header_line) + 100)

    rows = []
    for line in lines[header_line_idx + 1:]:
        stripped = line.strip()
        if not stripped or stripped.startswith("=") or stripped.startswith("---"):
            continue
        raw_values = []
        for k in range(len(col_positions) - 1):
            start = col_positions[k]
            end = col_positions[k + 1]
            val = line[start:end].strip() if start < len(line) else ""
            raw_values.append(val)
        if len(raw_values) >= 7:
            row = {
                "Designator": raw_values[0].strip('"').strip(),
                "Comment": raw_values[1].strip('"').strip(),
                "Layer": raw_values[2].strip('"').strip(),
                "Footprint": raw_values[3].strip('"').strip(),
                "Center-X(mm)": raw_values[4].strip('"').strip(),
                "Center-Y(mm)": raw_values[5].strip('"').strip(),
                "Rotation": raw_values[6].strip('"').strip(),
                "Description": raw_values[7].strip('"').strip() if len(raw_values) > 7 else "",
            }
            rows.append(row)

    return metadata, rows


def parse_csv_file(filepath):
    """Parse an Altium Pick and Place CSV file."""
    encoding = detect_encoding(filepath)
    rows = []
    with open(filepath, "r", encoding=encoding) as f:
        reader = csv.reader(f)
        csv_headers = None
        for line_idx, row in enumerate(reader):
            if not row or all(cell.strip() == "" for cell in row):
                continue
            if line_idx == 0 and "Designator" not in row:
                continue
            if "Designator" in row and "Center-X" in str(row):
                csv_headers = [h.strip() for h in row]
                continue
            if csv_headers and len(row) >= 7:
                entry = {}
                for i, h in enumerate(csv_headers[:len(row)]):
                    entry[h] = row[i].strip()
                rows.append(entry)
    return rows


# =============================================================================
# POCISAO Conversion Engine
# =============================================================================

POCISAO_COLUMNS = ["Designator", "Mid X", "Mid Y", "Layer", "Rotation"]


def convert_to_pocisao(parsed_rows):
    """Convert Pick and Place rows to PC POCISAO format."""
    converted = []
    layer_map = {
        "TopLayer": "Top", "Top": "Top",
        "BottomLayer": "Bottom", "Bottom": "Bottom",
    }
    for row in parsed_rows:
        x_val = row.get("Center-X(mm)", "0")
        y_val = row.get("Center-Y(mm)", "0")
        try:
            x_fmt = f"{float(x_val):.4f} mm"
        except (ValueError, TypeError):
            x_fmt = f"{x_val} mm"
        try:
            y_fmt = f"{float(y_val):.4f} mm"
        except (ValueError, TypeError):
            y_fmt = f"{y_val} mm"

        raw_layer = row.get("Layer", "Top").strip()
        layer = layer_map.get(raw_layer, raw_layer) or "Top"

        converted.append({
            "Designator": row.get("Designator", ""),
            "Mid X": x_fmt,
            "Mid Y": y_fmt,
            "Layer": layer,
            "Rotation": row.get("Rotation", "0"),
        })
    return converted


# =============================================================================
# BOM Conversion Engine
# =============================================================================

BOM_COLUMNS = ["Comment", "Designator", "Footprint", "JLCPCB Part #"]


def convert_to_bom(parsed_rows):
    """
    Convert Pick and Place rows to BOM format.
    Groups components by Comment (value field), case-insensitive.
    """
    groups = {}
    for row in parsed_rows:
        comment = row.get("Comment", "").strip()
        if not comment:
            continue
        key = comment.upper()  # case-insensitive grouping
        if key not in groups:
            groups[key] = {
                "designators": [],
                "description": row.get("Description", ""),
                "comment_orig": comment,
            }
        groups[key]["designators"].append(row.get("Designator", ""))

    result = []
    for key in sorted(groups.keys()):
        g = groups[key]
        # Sort designators naturally (C1, C2, C10, etc.)
        desigs = sorted(g["designators"], key=lambda d: (
            d[0] if d else "",
            int("".join(filter(str.isdigit, d))) if any(c.isdigit() for c in d) else 0
        ))
        result.append({
            "Comment": g["description"] or g["comment_orig"],
            "Designator": ", ".join(desigs),
            "Footprint": g["comment_orig"],
            "JLCPCB Part #": "",
        })
    return result


# =============================================================================
# Excel Export (shared)
# =============================================================================


def export_xlsx(rows, columns, output_path, col_widths=None):
    """Export rows to XLSX with given columns, styles, and widths."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for ri, row in enumerate(rows, 2):
        for ci, col_name in enumerate(columns, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col_name, ""))
            cell.alignment = data_align
            cell.border = thin_border

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            col_letter = chr(64 + i) if i <= 26 else "A"
            ws.column_dimensions[col_letter].width = w
    else:
        for ci in range(1, len(columns) + 1):
            ws.column_dimensions[chr(64 + ci)].width = 18

    wb.save(output_path)
    return output_path


# =============================================================================
# Donation / About Dialog
# =============================================================================

PAYPAL_DONATION_URL = "https://www.paypal.com/donate/?hosted_button_id=HW9QS2TXX9ZYQ"
PIX_KEY = "osmarjunioberaldo@hotmail.com"
DEV_URL = "https://github.com/DevOsmar"


def ensure_dependencies(parent=None, lang=None):
    """Check and auto-install openpyxl."""
    global openpyxl, HAS_OPENPYXL, Font, Alignment, Border, Side, PatternFill
    if HAS_OPENPYXL:
        return True

    def t(key, **kw):
        if lang is not None:
            return lang.t(key, **kw)
        d = LANGUAGES.get("pt_BR", {})
        val = d.get(key, key)
        if kw:
            try:
                val = val.format(**kw)
            except KeyError:
                pass
        return val

    install = messagebox.askyesno(
        t("dep_title"),
        t("dep_msg"),
        icon="question", parent=parent,
    )
    if not install:
        return False

    try:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        import importlib
        importlib.invalidate_caches()
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        HAS_OPENPYXL = True
        return True
    except Exception:
        messagebox.showerror(
            t("dep_install_error"),
            t("dep_install_msg"),
            parent=parent,
        )
        return False


# =============================================================================
# Reusable tab helpers
# =============================================================================

def sort_natural(items, col):
    """Sort items by column: numeric-aware for columns ending in digits."""
    def sort_key(item):
        val = (item.get(col) or "").strip()
        # Try natural sort: split into text + number parts
        parts = []
        buf = ""
        for ch in val:
            if ch.isdigit() != buf.isdigit() if buf else False:
                parts.append(buf)
                buf = ch
            else:
                buf += ch
        parts.append(buf)
        return [int(p) if p.isdigit() else p.lower() for p in parts]
    return sorted(items, key=sort_key)


# =============================================================================
# ToolTip (hover help)
# =============================================================================

class ToolTip:
    """Reusable tooltip that appears on hover over a widget."""

    def __init__(self, widget, text, delay=400):
        self.widget = widget
        self.text = text
        self.delay = delay
        self._tip_window = None
        self._after_id = None
        widget.bind("<Enter>", self._schedule, add="+")
        widget.bind("<Leave>", self._hide, add="+")
        widget.bind("<ButtonPress>", self._hide, add="+")

    def _schedule(self, event=None):
        self._hide()
        self._after_id = self.widget.after(self.delay, self._show)

    def _show(self):
        if self._tip_window:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
        self._tip_window = tw = Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tw.attributes("-topmost", True)
        label = Label(tw, text=self.text, justify="left",
                       background="#1E293B", foreground="#FFFFFF",
                       font=("Segoe UI", 8), padx=8, pady=4,
                       relief="flat", borderwidth=0)
        label.pack()

    def _hide(self, event=None):
        if self._after_id:
            self.widget.after_cancel(self._after_id)
            self._after_id = None
        if self._tip_window:
            self._tip_window.destroy()
            self._tip_window = None


# =============================================================================
# Tab 1: Pick Place -> POCISAO
# =============================================================================

class PocisaoTab:
    """Tab for converting Pick and Place -> POCISAO format."""

    def __init__(self, parent, app):
        self.app = app
        self.frame = ttk.Frame(parent, padding="10")
        self.frame.pack(fill="both", expand=True)

        self.input_filepath = StringVar()
        self.parsed_rows = []
        self.metadata = {}
        self.output_format = StringVar(value="xlsx")
        self.include_metadata = BooleanVar(value=True)
        self._edit_entry = None
        self._edit_item = None
        self._edit_col = None

        self._build()

    def _build(self):
        L = self.app.lang.t

        # Scrollable canvas
        self.canvas = Canvas(self.frame, borderwidth=0, highlightthickness=0)
        self.canvas_scroll = ttk.Scrollbar(self.frame, orient="vertical", command=self.canvas.yview)
        self.scroll_frame = ttk.Frame(self.canvas)
        self.scroll_frame.bind("<Configure>",
                               lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.configure(yscrollcommand=self.canvas_scroll.set)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.canvas_scroll.grid(row=0, column=1, sticky="ns")
        self.frame.grid_rowconfigure(0, weight=1)
        self.frame.grid_columnconfigure(0, weight=1)

        # Stretch scroll_frame to canvas width on resize
        self._canvas_window = self.canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        self.canvas.bind("<Configure>",
                         lambda e: self.canvas.itemconfig(self._canvas_window, width=e.width))

        # Mouse wheel scrolling (bound to the canvas widget, not globally)
        def _on_mousewheel(event):
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self.canvas.bind("<MouseWheel>", _on_mousewheel)

        sf = self.scroll_frame

        # Header
        header = ttk.Frame(sf)
        header.pack(fill="x", pady=(0, 10))
        ttk.Label(header, text=L("pocisao_header"),
                  font=("Segoe UI", 12, "bold")).pack(anchor="w")
        ttk.Label(header,
                  text=L("pocisao_desc"),
                  font=("Segoe UI", 9), foreground=self.app.COLORS["text_secondary"]).pack(anchor="w", pady=(2, 0))

        # File selection
        sel = ttk.LabelFrame(sf, text=L("input_file"), padding="10")
        sel.pack(fill="x", pady=(0, 10))
        row1 = ttk.Frame(sel)
        row1.pack(fill="x", pady=(3, 0))
        ttk.Label(row1, text=L("input_file_label")).pack(side="left")
        self.file_entry = ttk.Entry(row1, textvariable=self.input_filepath, font=("Consolas", 9))
        self.file_entry.pack(side="left", fill="x", expand=True, padx=(8, 8))
        btn_sel = ttk.Button(row1, text=L("btn_select"), command=self.select_file,
                           style="Outline.TButton")
        btn_sel.pack(side="left")
        ToolTip(btn_sel, L("tip_select_file"))
        self.file_info = ttk.Label(sel, text="", foreground=self.app.COLORS["text_secondary"])
        self.file_info.pack(anchor="w", pady=(6, 0))

        # Preview
        prev = ttk.LabelFrame(sf, text=L("preview_data"), padding="8")
        prev.pack(fill="both", expand=True, pady=(0, 10))

        self.tree = ttk.Treeview(prev, columns=POCISAO_COLUMNS, show="headings",
                                  height=8, selectmode="browse")
        widths = {"Designator": 120, "Mid X": 120, "Mid Y": 120, "Layer": 100, "Rotation": 80}
        for col in POCISAO_COLUMNS:
            self.tree.heading(col, text=col,
                              command=lambda c=col: self._sort(c))
            self.tree.column(col, width=widths.get(col, 100), anchor="center", minwidth=60)

        vsb = ttk.Scrollbar(prev, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(prev, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        prev.grid_rowconfigure(0, weight=1)
        prev.grid_columnconfigure(0, weight=1)
        self.status = ttk.Label(prev, text=L("no_file_loaded"), foreground=self.app.COLORS["text_muted"])
        self.status.grid(row=2, column=0, sticky="w", pady=(2, 0))

        # Treeview bindings for editing
        self.tree.bind("<Double-1>", self._start_edit)
        self.tree.bind("<Button-3>", self._on_right_click)
        self.tree.bind("<Delete>", lambda e: self._delete_row())

        # Action buttons
        actions = ttk.Frame(prev)
        actions.grid(row=3, column=0, sticky="w", pady=(8, 0))
        btn_add = ttk.Button(actions, text=L("btn_add_row"), command=self._add_row,
                            style="Small.TButton")
        btn_add.pack(side="left", padx=(0, 8))
        ToolTip(btn_add, L("tip_add_row"))
        btn_del = ttk.Button(actions, text=L("btn_delete_row"), command=self._delete_row,
                            style="Small.TButton")
        btn_del.pack(side="left")
        ToolTip(btn_del, L("tip_delete_row"))

        # Output (ao final)
        out = ttk.LabelFrame(sf, text=L("output_options"), padding="10")
        out.pack(fill="x", pady=(0, 10))

        fmt_row = ttk.Frame(out)
        fmt_row.pack(fill="x", pady=(0, 5))
        ttk.Label(fmt_row, text=L("format_label")).pack(side="left")
        ttk.Radiobutton(fmt_row, text=L("format_xlsx"), variable=self.output_format,
                        value="xlsx").pack(side="left", padx=(8, 0))
        ttk.Radiobutton(fmt_row, text=L("format_xml"), variable=self.output_format,
                        value="xml").pack(side="left", padx=(8, 0))

        opt_row = ttk.Frame(out)
        opt_row.pack(fill="x")
        ttk.Checkbutton(opt_row, text=L("include_metadata"),
                        variable=self.include_metadata).pack(side="left")

        # Action row for progress bar and convert button
        btn_row = ttk.Frame(out)
        btn_row.pack(fill="x", pady=(5, 0))

        self.progress = ttk.Progressbar(btn_row, mode="indeterminate", length=200)
        self.progress.pack(side="right", padx=(0, 10))
        self.progress.pack_forget()

        btn_conv = ttk.Button(btn_row, text=L("btn_convert"), command=self.convert,
                            style="Accent.TButton")
        btn_conv.pack(side="right")
        ToolTip(btn_conv, L("tip_convert"))


    def select_file(self):
        L = self.app.lang.t
        f = filedialog.askopenfilename(
            title=L("select_pnp_title"),
            filetypes=[(L("supported_files"), "*.txt *.csv *.xlsx *.xml"),
                       ("Pick and Place TXT", "*.txt"), ("CSV", "*.csv"),
                       ("Excel", "*.xlsx"), ("XML", "*.xml"), (L("all_files"), "*.*")])
        if f:
            self.load(f)

    def load(self, filepath):
        self.input_filepath.set(filepath)
        try:
            ext = os.path.splitext(filepath)[1].lower()
            if ext == ".txt":
                md, pr = parse_pick_place_txt(filepath)
                self.metadata = md
                self.parsed_rows = convert_to_pocisao(pr)
            elif ext == ".csv":
                pr = parse_csv_file(filepath)
                self.metadata = {}
                self.parsed_rows = convert_to_pocisao(pr)
            elif ext == ".xlsx":
                wb = openpyxl.load_workbook(filepath, data_only=True)
                ws = wb.active
                all_rows = []
                for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
                    vals = [str(v).strip() if v is not None else "" for v in row]
                    if ri == 1:
                        hdrs = vals
                    elif any(v for v in vals):
                        all_rows.append({hdrs[i]: vals[i] if i < len(vals) else "" for i in range(len(hdrs))})
                wb.close()
                if all_rows and "Center-X(mm)" in all_rows[0]:
                    self.parsed_rows = convert_to_pocisao(all_rows)
                else:
                    self.parsed_rows = all_rows
                self.metadata = {}
            elif ext == ".xml":
                self.parsed_rows = self._parse_xml(filepath)
                self.metadata = {}
            else:
                raise ValueError(f"Formato não suportado: {ext}")

            self._update_preview()
            L = self.app.lang.t
            units = self.metadata.get("Units", "mm")
            self.file_info.config(
                text=L("file_info_txt",
                       file=os.path.basename(filepath),
                       count=len(self.parsed_rows),
                       components=L("components"),
                       units_label=L("units"),
                       units=units),
                foreground=self.app.COLORS["success"])
            self.app.set_status(L("status_file_loaded", file=os.path.basename(filepath)))
        except Exception as e:
            messagebox.showerror(self.app.lang.t("error_load"), str(e))
            self.app.set_status(L("status_error", msg=str(e)), is_error=True)
            self.parsed_rows = []
            self.metadata = {}
            self._update_preview()

    def _parse_xml(self, filepath):
        tree = ET.parse(filepath)
        root = tree.getroot()
        rows = []
        for comp in root.findall(".//Component"):
            row = {}
            for child in comp:
                row[child.tag.replace("_", " ")] = child.text or ""
            if row.get("Designator"):
                rows.append(row)
        return rows

    def _update_preview(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        C = self.app.COLORS
        self.tree.tag_configure("even", background=C["row_even"])
        self.tree.tag_configure("odd", background=C["row_odd"])
        for i, r in enumerate(self.parsed_rows):
            tag = "even" if i % 2 == 0 else "odd"
            self.tree.insert("", "end", tags=(tag,), values=(
                r.get("Designator", ""), r.get("Mid X", ""),
                r.get("Mid Y", ""), r.get("Layer", ""),
                r.get("Rotation", "")))
        L = self.app.lang.t
        cnt = len(self.parsed_rows)
        if cnt:
            self.status.config(text=f"{cnt} {L('components')}", foreground=self.app.COLORS["text"])
        else:
            self.status.config(text=L("no_data"), foreground=self.app.COLORS["text_muted"])

    def _sort(self, col):
        items = [{c: self.tree.set(k, c) for c in POCISAO_COLUMNS}
                 for k in self.tree.get_children("")]
        items = sort_natural(items, col)
        self.tree.delete(*self.tree.get_children())
        for i, item in enumerate(items):
            tag = "even" if i % 2 == 0 else "odd"
            self.tree.insert("", "end", tags=(tag,), values=[item.get(c, "") for c in POCISAO_COLUMNS])

    # ---- Editing, Adding, Deleting ----

    def _add_row(self):
        """Add a new blank row to the data."""
        blank = {c: "" for c in POCISAO_COLUMNS}
        self.parsed_rows.append(blank)
        self.tree.insert("", "end", values=[blank.get(c, "") for c in POCISAO_COLUMNS])
        self.status.config(text=f"{len(self.parsed_rows)} {self.app.lang.t('components')}", foreground=self.app.COLORS["text"])

    def _delete_row(self):
        """Delete the selected row from the data."""
        sel = self.tree.selection()
        if not sel:
            return
        item_id = sel[0]
        idx = self.tree.index(item_id)
        self.tree.delete(item_id)
        if 0 <= idx < len(self.parsed_rows):
            del self.parsed_rows[idx]
        cnt = len(self.parsed_rows)
        L = self.app.lang.t
        if cnt:
            self.status.config(text=f"{cnt} {L('components')}", foreground=self.app.COLORS["text"])
        else:
            self.status.config(text=L("no_data"), foreground=self.app.COLORS["text_muted"])

    def _start_edit(self, event):
        """Start inline editing on double-click."""
        if self._edit_entry is not None:
            self._finish_edit()

        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return

        col_id = self.tree.identify_column(event.x)   # e.g. "#1"
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return

        # Map column index to column name
        col_idx = int(col_id.replace("#", "")) - 1
        if col_idx < 0 or col_idx >= len(POCISAO_COLUMNS):
            return
        col_name = POCISAO_COLUMNS[col_idx]

        # Get cell bounding box
        bbox = self.tree.bbox(item_id, col_id)
        if not bbox:
            return
        x, y, w, h = bbox

        current_val = self.tree.set(item_id, col_name)

        self._edit_item = item_id
        self._edit_col = col_name

        # Use Combobox for Layer column (Top/Bottom), plain Entry for others
        if col_name == "Layer":
            entry = ttk.Combobox(self.tree, values=["Top", "Bottom"],
                                  state="readonly", font=("Consolas", 9))
            entry.set(current_val if current_val in ("Top", "Bottom") else "Top")
            entry.place(x=x, y=y, width=w, height=h)
            entry.focus_set()
            entry.bind("<<ComboboxSelected>>", self._finish_edit)
            entry.bind("<Escape>", self._cancel_edit)
            entry.bind("<FocusOut>", self._finish_edit)
        else:
            entry = ttk.Entry(self.tree, font=("Consolas", 9))
            entry.place(x=x, y=y, width=w, height=h)
            entry.insert(0, current_val)
            entry.select_range(0, "end")
            entry.focus_set()
            entry.bind("<Return>", self._finish_edit)
            entry.bind("<Escape>", self._cancel_edit)
            entry.bind("<FocusOut>", self._finish_edit)

        self._edit_entry = entry

    def _finish_edit(self, event=None):
        """Save the edited value and remove the entry widget."""
        if self._edit_entry is None:
            return
        new_val = self._edit_entry.get()
        item_id = self._edit_item
        col_name = self._edit_col

        # Update tree display
        self.tree.set(item_id, col_name, new_val)

        # Update data source
        idx = self.tree.index(item_id)
        if 0 <= idx < len(self.parsed_rows):
            self.parsed_rows[idx][col_name] = new_val

        self._destroy_edit()

    def _cancel_edit(self, event=None):
        """Cancel editing without saving."""
        self._destroy_edit()

    def _destroy_edit(self):
        """Destroy the edit entry widget and reset state."""
        if self._edit_entry:
            try:
                self._edit_entry.destroy()
            except TclError:
                pass
        self._edit_entry = None
        self._edit_item = None
        self._edit_col = None

    def _on_right_click(self, event):
        """Show right-click context menu."""
        item_id = self.tree.identify_row(event.y)
        if item_id:
            self.tree.selection_set(item_id)

        L = self.app.lang.t
        menu = Menu(self.frame, tearoff=0)
        menu.add_command(label=L("menu_edit"), command=lambda: self._start_edit(
            type("ev", (object,), {"x": event.x, "y": event.y})()))
        menu.add_separator()
        menu.add_command(label=L("btn_add_row"), command=self._add_row)
        menu.add_command(label=L("btn_delete_row"), command=self._delete_row)
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def convert(self):
        L = self.app.lang.t
        if not self.parsed_rows:
            messagebox.showwarning(L("warn_no_data"), L("warn_load_first"))
            return

        fmt = self.output_format.get()
        ext = ".xlsx" if fmt == "xlsx" else ".xml"
        ft = [("Excel XLSX", "*.xlsx"), (L("all_files"), "*.*")] if fmt == "xlsx" else [("XML", "*.xml"), (L("all_files"), "*.*")]
        default_name = f"JLCSMT Convert New{ext}"

        out = filedialog.asksaveasfilename(title=L("dialog_save_as"), defaultextension=ext,
                                            initialfile=default_name, filetypes=ft)
        if not out:
            return

        self.progress.pack(side="right", padx=(0, 10))
        self.progress.start()
        self.frame.update_idletasks()

        try:
            if fmt == "xlsx":
                export_xlsx(self.parsed_rows, POCISAO_COLUMNS, out, [15, 14, 14, 10, 10])
            else:
                root = ET.Element("PCB_PickAndPlace")
                if self.metadata and self.include_metadata.get():
                    meta = ET.SubElement(root, "Metadata")
                    for k, v in self.metadata.items():
                        ET.SubElement(meta, k.replace(" ", "_")).text = str(v)
                comps = ET.SubElement(root, "Components")
                for r in self.parsed_rows:
                    c = ET.SubElement(comps, "Component")
                    for col in POCISAO_COLUMNS:
                        ET.SubElement(c, col.replace(" ", "_")).text = str(r.get(col, ""))
                dom = minidom.parseString(ET.tostring(root, encoding="unicode").encode("utf-8"))
                with open(out, "wb") as f:
                    f.write(dom.toprettyxml(indent="  ", encoding="utf-8"))

            self.progress.stop()
            self.progress.pack_forget()
            L = self.app.lang.t
            self.app.set_status(L("status_saved", file=os.path.basename(out)))
            msg = L("file_saved_msg", path=out, count=len(self.parsed_rows), components=L("components"))
            if messagebox.askyesno(L("dialog_complete"), msg + "\n\n" + L("ask_open_file")):
                self.app._open_file(out)
        except Exception as e:
            self.progress.stop()
            self.progress.pack_forget()
            self.app.set_status(L("status_error", msg=str(e)), is_error=True)
            messagebox.showerror(self.app.lang.t("error_title"), f"{self.app.lang.t('error_convert')}\n{str(e)}")


# =============================================================================
# Tab 2: Pick Place -> BOM
# =============================================================================

class BomTab:
    """Tab for converting Pick and Place CSV -> BOM format."""

    def __init__(self, parent, app):
        self.app = app
        self.frame = ttk.Frame(parent, padding="10")
        self.frame.pack(fill="both", expand=True)

        self.input_filepath = StringVar()
        self.parsed_rows = []
        self.raw_rows = []
        self._edit_entry = None
        self._edit_item = None
        self._edit_col = None

        self._build()

    def _build(self):
        L = self.app.lang.t

        # Scrollable canvas
        self.canvas = Canvas(self.frame, borderwidth=0, highlightthickness=0)
        self.canvas_scroll = ttk.Scrollbar(self.frame, orient="vertical", command=self.canvas.yview)
        self.scroll_frame = ttk.Frame(self.canvas)
        self.scroll_frame.bind("<Configure>",
                               lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.configure(yscrollcommand=self.canvas_scroll.set)
        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.canvas_scroll.grid(row=0, column=1, sticky="ns")
        self.frame.grid_rowconfigure(0, weight=1)
        self.frame.grid_columnconfigure(0, weight=1)

        # Stretch scroll_frame to canvas width on resize
        self._canvas_window = self.canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        self.canvas.bind("<Configure>",
                         lambda e: self.canvas.itemconfig(self._canvas_window, width=e.width))

        # Mouse wheel scrolling (bound to the canvas widget, not globally)
        def _on_mousewheel(event):
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self.canvas.bind("<MouseWheel>", _on_mousewheel)

        sf = self.scroll_frame

        # Header
        header = ttk.Frame(sf)
        header.pack(fill="x", pady=(0, 10))
        ttk.Label(header, text=L("bom_header"),
                  font=("Segoe UI", 12, "bold")).pack(anchor="w")
        ttk.Label(header,
                  text=L("bom_desc"),
                  font=("Segoe UI", 9), foreground=self.app.COLORS["text_secondary"]).pack(anchor="w", pady=(2, 0))

        # File selection
        sel = ttk.LabelFrame(sf, text=L("input_file_csv"), padding="10")
        sel.pack(fill="x", pady=(0, 10))
        row1 = ttk.Frame(sel)
        row1.pack(fill="x", pady=(3, 0))
        ttk.Label(row1, text=L("input_file_label")).pack(side="left")
        self.file_entry = ttk.Entry(row1, textvariable=self.input_filepath, font=("Consolas", 9))
        self.file_entry.pack(side="left", fill="x", expand=True, padx=(8, 8))
        btn_sel = ttk.Button(row1, text=L("btn_select"), command=self.select_file,
                           style="Outline.TButton")
        btn_sel.pack(side="left")
        ToolTip(btn_sel, L("tip_select_file"))
        self.file_info = ttk.Label(sel, text="", foreground=self.app.COLORS["text_secondary"])
        self.file_info.pack(anchor="w", pady=(6, 0))

        # Preview
        prev = ttk.LabelFrame(sf, text=L("preview_bom"), padding="8")
        prev.pack(fill="both", expand=True, pady=(0, 10))

        self.tree = ttk.Treeview(prev, columns=BOM_COLUMNS, show="headings",
                                  height=8, selectmode="browse")
        widths = {"Comment": 300, "Designator": 180, "Footprint": 150, "JLCPCB Part #": 120}
        for col in BOM_COLUMNS:
            self.tree.heading(col, text=col, command=lambda c=col: self._sort(c))
            self.tree.column(col, width=widths.get(col, 150), anchor="center", minwidth=60)

        vsb = ttk.Scrollbar(prev, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(prev, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        prev.grid_rowconfigure(0, weight=1)
        prev.grid_columnconfigure(0, weight=1)
        self.status = ttk.Label(prev, text=L("no_file_loaded"), foreground=self.app.COLORS["text_muted"])
        self.status.grid(row=2, column=0, sticky="w", pady=(2, 0))

        # Treeview bindings for editing
        self.tree.bind("<Double-1>", self._start_edit)
        self.tree.bind("<Button-3>", self._on_right_click)
        self.tree.bind("<Delete>", lambda e: self._delete_row())

        # Action buttons
        actions = ttk.Frame(prev)
        actions.grid(row=3, column=0, sticky="w", pady=(8, 0))
        btn_add = ttk.Button(actions, text=L("btn_add_row"), command=self._add_row,
                            style="Small.TButton")
        btn_add.pack(side="left", padx=(0, 8))
        ToolTip(btn_add, L("tip_add_row"))
        btn_del = ttk.Button(actions, text=L("btn_delete_row"), command=self._delete_row,
                            style="Small.TButton")
        btn_del.pack(side="left")
        ToolTip(btn_del, L("tip_delete_row"))

        # Search buttons em linha única
        search_frame = ttk.Frame(prev)
        search_frame.grid(row=4, column=0, sticky="w", pady=(4, 0))
        ttk.Label(search_frame, text=L("btn_search"), font=("Segoe UI", 8)).pack(anchor="w")
        search_row = ttk.Frame(search_frame)
        search_row.pack(fill="x")
        btn_octo = ttk.Button(search_row, text=L("search_octopart"),
                              command=lambda: self._search_component("octopart"),
                              style="Small.TButton")
        btn_octo.pack(side="left", padx=(0, 4))
        ToolTip(btn_octo, L("tip_search_octopart"))
        btn_snap = ttk.Button(search_row, text=L("search_snapmagic"),
                              command=lambda: self._search_component("snapmagic"),
                              style="Small.TButton")
        btn_snap.pack(side="left", padx=(0, 4))
        ToolTip(btn_snap, L("tip_search_snapmagic"))
        btn_mouser = ttk.Button(search_row, text=L("search_mouser"),
                                command=lambda: self._search_component("mouser"),
                                style="Small.TButton")
        btn_mouser.pack(side="left", padx=(0, 4))
        ToolTip(btn_mouser, L("tip_search_mouser"))
        btn_jlcpcb = ttk.Button(search_row, text=L("search_jlcpcb"),
                                command=lambda: self._search_component("jlcpcb"),
                                style="Small.TButton")
        btn_jlcpcb.pack(side="left", padx=(0, 4))
        ToolTip(btn_jlcpcb, L("tip_search_jlcpcb"))
        btn_digikey = ttk.Button(search_row, text=L("search_digikey"),
                                 command=lambda: self._search_component("digikey"),
                                 style="Small.TButton")
        btn_digikey.pack(side="left", padx=(0, 4))
        ToolTip(btn_digikey, L("tip_search_digikey"))
        btn_snapeda = ttk.Button(search_row, text=L("search_snapeda"),
                                 command=lambda: self._search_component("snapeda"),
                                 style="Small.TButton")
        btn_snapeda.pack(side="left", padx=(0, 4))
        ToolTip(btn_snapeda, L("tip_search_snapeda"))
        btn_lcsc = ttk.Button(search_row, text=L("search_lcsc"),
                              command=lambda: self._search_component("lcsc"),
                              style="Small.TButton")
        btn_lcsc.pack(side="left", padx=(0, 4))
        ToolTip(btn_lcsc, L("tip_search_lcsc"))

        # Output (ao final) - BomTab
        out = ttk.LabelFrame(sf, text=L("output_options"), padding="10")
        out.pack(fill="x", pady=(0, 10))

        btn_row = ttk.Frame(out)
        btn_row.pack(fill="x", pady=(3, 0))

        self.progress = ttk.Progressbar(btn_row, mode="indeterminate", length=200)
        self.progress.pack(side="right", padx=(0, 10))
        self.progress.pack_forget()

        btn_conv = ttk.Button(btn_row, text=L("btn_convert"), command=self.convert,
                            style="Accent.TButton")
        btn_conv.pack(side="right")
        ToolTip(btn_conv, L("tip_convert"))

    def select_file(self):
        L = self.app.lang.t
        f = filedialog.askopenfilename(
            title=L("select_csv_title"),
            filetypes=[("CSV Pick and Place", "*.csv"), ("Pick and Place TXT", "*.txt"),
                       (L("all_files"), "*.*")])
        if f:
            self.load(f)

    def load(self, filepath):
        self.input_filepath.set(filepath)
        try:
            ext = os.path.splitext(filepath)[1].lower()
            if ext == ".csv":
                self.raw_rows = parse_csv_file(filepath)
            elif ext == ".txt":
                _, self.raw_rows = parse_pick_place_txt(filepath)
            else:
                raise ValueError(f"{self.app.lang.t('unsupported_format')}: {ext}")

            self.parsed_rows = convert_to_bom(self.raw_rows)
            self._update_preview()
            L = self.app.lang.t
            self.file_info.config(
                text=L("file_info_bom",
                       file=os.path.basename(filepath),
                       raw=len(self.raw_rows),
                       count=len(self.parsed_rows)),
                foreground=self.app.COLORS["success"])
            self.app.set_status(L("status_file_loaded", file=os.path.basename(filepath)))
        except Exception as e:
            messagebox.showerror(self.app.lang.t("error_load"), str(e))
            self.app.set_status(L("status_error", msg=str(e)), is_error=True)
            self.parsed_rows = []
            self.raw_rows = []
            self._update_preview()

    def _update_preview(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        C = self.app.COLORS
        self.tree.tag_configure("even", background=C["row_even"])
        self.tree.tag_configure("odd", background=C["row_odd"])
        for i, r in enumerate(self.parsed_rows):
            tag = "even" if i % 2 == 0 else "odd"
            self.tree.insert("", "end", tags=(tag,), values=(
                r.get("Comment", ""), r.get("Designator", ""),
                r.get("Footprint", ""), r.get("JLCPCB Part #", "")))
        L = self.app.lang.t
        cnt = len(self.parsed_rows)
        if cnt:
            self.status.config(text=f"{cnt} {L('component_groups')}", foreground=self.app.COLORS["text"])
        else:
            self.status.config(text=L("no_data"), foreground=self.app.COLORS["text_muted"])

    def _sort(self, col):
        items = [{c: self.tree.set(k, c) for c in BOM_COLUMNS} for k in self.tree.get_children("")]
        items = sort_natural(items, col)
        self.tree.delete(*self.tree.get_children())
        for i, item in enumerate(items):
            tag = "even" if i % 2 == 0 else "odd"
            self.tree.insert("", "end", tags=(tag,), values=[item.get(c, "") for c in BOM_COLUMNS])

    # ---- Editing, Adding, Deleting ----

    def _add_row(self):
        """Add a new blank row to the data."""
        blank = {c: "" for c in BOM_COLUMNS}
        self.parsed_rows.append(blank)
        self.tree.insert("", "end", values=[blank.get(c, "") for c in BOM_COLUMNS])
        L = self.app.lang.t
        self.status.config(text=f"{len(self.parsed_rows)} {L('component_groups')}", foreground=self.app.COLORS["text"])

    def _delete_row(self):
        """Delete the selected row from the data."""
        sel = self.tree.selection()
        if not sel:
            return
        item_id = sel[0]
        idx = self.tree.index(item_id)
        self.tree.delete(item_id)
        if 0 <= idx < len(self.parsed_rows):
            del self.parsed_rows[idx]
        L = self.app.lang.t
        cnt = len(self.parsed_rows)
        if cnt:
            self.status.config(text=f"{cnt} {L('component_groups')}", foreground=self.app.COLORS["text"])
        else:
            self.status.config(text=L("no_data"), foreground=self.app.COLORS["text_muted"])

    def _start_edit(self, event):
        """Start inline editing on double-click."""
        if self._edit_entry is not None:
            self._finish_edit()

        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return

        col_id = self.tree.identify_column(event.x)
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return

        col_idx = int(col_id.replace("#", "")) - 1
        if col_idx < 0 or col_idx >= len(BOM_COLUMNS):
            return
        col_name = BOM_COLUMNS[col_idx]

        bbox = self.tree.bbox(item_id, col_id)
        if not bbox:
            return
        x, y, w, h = bbox

        current_val = self.tree.set(item_id, col_name)

        self._edit_item = item_id
        self._edit_col = col_name

        entry = ttk.Entry(self.tree, font=("Consolas", 9))
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, current_val)
        entry.select_range(0, "end")
        entry.focus_set()

        entry.bind("<Return>", self._finish_edit)
        entry.bind("<Escape>", self._cancel_edit)
        entry.bind("<FocusOut>", self._finish_edit)

        self._edit_entry = entry

    def _finish_edit(self, event=None):
        """Save the edited value and remove the entry widget."""
        if self._edit_entry is None:
            return
        new_val = self._edit_entry.get()
        item_id = self._edit_item
        col_name = self._edit_col

        self.tree.set(item_id, col_name, new_val)

        idx = self.tree.index(item_id)
        if 0 <= idx < len(self.parsed_rows):
            self.parsed_rows[idx][col_name] = new_val

        self._destroy_edit()

    def _cancel_edit(self, event=None):
        """Cancel editing without saving."""
        self._destroy_edit()

    def _destroy_edit(self):
        """Destroy the edit entry widget and reset state."""
        if self._edit_entry:
            try:
                self._edit_entry.destroy()
            except TclError:
                pass
        self._edit_entry = None
        self._edit_item = None
        self._edit_col = None

    def _on_right_click(self, event):
        """Show right-click context menu."""
        item_id = self.tree.identify_row(event.y)
        if item_id:
            self.tree.selection_set(item_id)

        L = self.app.lang.t
        menu = Menu(self.frame, tearoff=0)
        menu.add_command(label=L("menu_edit"), command=lambda: self._start_edit(
            type("ev", (object,), {"x": event.x, "y": event.y})()))
        menu.add_separator()
        menu.add_command(label=L("btn_add_row"), command=self._add_row)
        menu.add_command(label=L("btn_delete_row"), command=self._delete_row)
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _search_component(self, site):
        """Open browser to search for the selected component on the given site."""
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning(
                self.app.lang.t("btn_search"),
                self.app.lang.t("search_no_selection"))
            return

        item_id = sel[0]
        comment = self.tree.set(item_id, "Comment")
        footprint = self.tree.set(item_id, "Footprint")
        designator = self.tree.set(item_id, "Designator")

        # Build the search query from the most meaningful fields
        query_parts = [p for p in [comment, footprint] if p.strip()]
        query = " ".join(query_parts) if query_parts else designator

        if not query.strip():
            messagebox.showwarning(
                self.app.lang.t("btn_search"),
                self.app.lang.t("search_no_selection"))
            return

        encoded = quote(query)

        urls = {
            "octopart": f"https://octopart.com/search?q={encoded}",
            "snapmagic": f"https://www.snapmagic.com/search?q={encoded}",
            "mouser": f"https://br.mouser.com/c/?q={encoded}",
            "jlcpcb": f"https://jlcpcb.com/parts/componentSearch?searchTxt={encoded}",
            "digikey": f"https://www.digikey.com.br/pt/products/result?s={encoded}",
            "snapeda": f"https://www.snapeda.com/search/?q={encoded}",
            "lcsc": f"https://www.lcsc.com/search?q={encoded}",
        }

        url = urls.get(site)
        if url:
            webbrowser.open(url)

    def convert(self):
        L = self.app.lang.t
        if not self.parsed_rows:
            messagebox.showwarning(L("warn_no_data"), L("warn_load_first"))
            return

        out = filedialog.asksaveasfilename(
            title=L("dialog_save_bom"),
            defaultextension=".xlsx",
            initialfile="Bom Convert New.xlsx",
            filetypes=[("Excel XLSX", "*.xlsx"), (L("all_files"), "*.*")])
        if not out:
            return

        self.progress.pack(side="right", padx=(0, 10))
        self.progress.start()
        self.frame.update_idletasks()

        try:
            export_xlsx(self.parsed_rows, BOM_COLUMNS, out, [40, 22, 18, 18])
            self.progress.stop()
            self.progress.pack_forget()
            self.app.set_status(L("status_saved", file=os.path.basename(out)))
            msg = L("bom_saved_msg", path=out, count=len(self.parsed_rows), groups=L("component_groups"))
            if messagebox.askyesno(L("dialog_complete"), msg + "\n\n" + L("ask_open_file")):
                self.app._open_file(out)
        except Exception as e:
            self.progress.stop()
            self.progress.pack_forget()
            self.app.set_status(L("status_error", msg=str(e)), is_error=True)
            messagebox.showerror(self.app.lang.t("error_title"), f"{self.app.lang.t('error_convert')}\n{str(e)}")


# =============================================================================
# Main Application
# =============================================================================

class ConverterApp:
    """Main application window with tabs."""

    # --- Color Palette (system default / neutral tones) ---
    COLORS = {
        "primary": "SystemButtonText",
        "primary_hover": "SystemButtonFace",
        "primary_light": "SystemButtonFace",
        "success": "#006600",
        "success_light": "#FFFFFF",
        "danger": "SystemHighlight",
        "danger_light": "SystemButtonFace",
        "warning": "SystemHighlightText",
        "bg": "SystemButtonFace",
        "surface": "SystemButtonFace",
        "text": "SystemWindowText",
        "text_secondary": "SystemGrayText",
        "text_muted": "SystemGrayText",
        "border": "SystemButtonFace",
        "header_bg": "SystemActiveCaption",
        "header_fg": "SystemButtonFace",
        "row_even": "SystemWindow",
        "row_odd": "SystemWindow",
        "selection_bg": "SystemHighlight",
        "selection_fg": "SystemHighlightText",
        "tab_bg": "SystemButtonFace",
        "tab_active": "SystemButtonFace",
    }

    def __init__(self, root):
        self.root = root
        self.lang = Lang()
        self.root.title(self.lang.t("app_title"))
        self.root.geometry("1100x760")
        self.root.minsize(850, 600)

        try:
            self.root.iconbitmap(default="")
        except TclError:
            pass

        self._build_menu()
        self._build_ui()

        self.root.update_idletasks()
        w, h = 950, 680
        x = (self.root.winfo_screenwidth() - w) // 2
        y = (self.root.winfo_screenheight() - h) // 2
        self.root.geometry(f"+{x}+{y}")

    def _build_menu(self):
        menubar = Menu(self.root)
        self.root.config(menu=menubar)

        # File menu
        file_menu = Menu(menubar, tearoff=0)
        file_menu.add_command(label=self.lang.t("menu_exit"), command=self.root.quit, accelerator="Ctrl+Q")
        menubar.add_cascade(label=self.lang.t("menu_file"), menu=file_menu)

        # Language menu
        lang_menu = Menu(menubar, tearoff=0)
        lang_menu.add_command(label=self.lang.t("lang_pt"), command=lambda: self._switch_lang("pt_BR"))
        lang_menu.add_command(label=self.lang.t("lang_en"), command=lambda: self._switch_lang("en_US"))
        menubar.add_cascade(label=self.lang.t("menu_language"), menu=lang_menu)

        # Help menu
        help_menu = Menu(menubar, tearoff=0)
        help_menu.add_command(label=self.lang.t("menu_about"), command=self.show_about)
        menubar.add_cascade(label=self.lang.t("menu_help"), menu=help_menu)
        self.root.bind("<Control-q>", lambda e: self.root.quit())
        self.root.bind("<Control-l>", lambda e: self._cycle_language())
        self.root.bind("<Control-L>", lambda e: self._cycle_language())

    def _cycle_language(self):
        """Toggle between pt_BR and en_US."""
        next_lang = "en_US" if self.lang.lang == "pt_BR" else "pt_BR"
        self._switch_lang(next_lang)

    def _switch_lang(self, lang_code):
        """Switch application language."""
        self.lang.switch_to(lang_code)
        self._rebuild_ui()

    def _rebuild_ui(self):
        """Rebuild the entire UI when language changes."""
        self.root.title(self.lang.t("app_title"))
        for widget in self.root.winfo_children():
            widget.destroy()
        self._build_menu()
        self._build_ui()

    def _build_ui(self):
        main = ttk.Frame(self.root, padding="5")
        main.pack(fill="both", expand=True)

        # Header
        header = ttk.Frame(main)
        header.pack(fill="x", pady=(0, 8))
        ttk.Label(header, text="\U0001f50c PCB Pick and Place Converter",
                  font=("Segoe UI", 14, "bold"),
                  foreground=self.COLORS["text"]).pack(anchor="w")

        # Notebook (tabs)
        self.notebook = ttk.Notebook(main)
        self.notebook.pack(fill="both", expand=True)

        # Tab 1 - POCISAO
        self.tab1 = PocisaoTab(self.notebook, self)
        self.notebook.add(self.tab1.frame, text=self.lang.t("tab_pocisao"))

        # Tab 2 - BOM
        self.tab2 = BomTab(self.notebook, self)
        self.notebook.add(self.tab2.frame, text=self.lang.t("tab_bom"))

        # Status bar
        status_frame = ttk.Frame(main, relief="sunken", padding=(4, 2))
        status_frame.pack(fill="x", pady=(2, 0))
        self.status_bar = ttk.Label(status_frame, text=f"  {self.lang.t('status_ready')}",
                                     font=("Segoe UI", 8),
                                     foreground=self.COLORS["text_secondary"])
        self.status_bar.pack(side="left")

    def set_status(self, text, is_error=False):
        """Update the status bar text."""
        fg = self.COLORS["danger"] if is_error else self.COLORS["text_secondary"]
        self.status_bar.config(text=f"  {text}", foreground=fg)

    def _open_file(self, filepath):
        try:
            if sys.platform == "win32":
                os.startfile(filepath)
            elif sys.platform == "darwin":
                subprocess.call(["open", filepath])
            else:
                subprocess.call(["xdg-open", filepath])
        except Exception as e:
            messagebox.showwarning(self.lang.t("open_file_error"), str(e))

    def show_about(self):
        dialog = Toplevel(self.root)
        dialog.title(self.lang.t("about_title"))
        dialog.geometry("480x520")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.update_idletasks()
        px = self.root.winfo_x() + (self.root.winfo_width() - 480) // 2
        py = self.root.winfo_y() + (self.root.winfo_height() - 520) // 2
        dialog.geometry(f"+{px}+{py}")

        C = self.COLORS
        bg = C["bg"]
        dialog.configure(bg=bg)
        frame = Frame(dialog, bg=bg, padx=25, pady=25)
        frame.pack(fill="both", expand=True)

        Label(frame, text="\U0001f50c PCB Pick and Place Converter",
              font=("Segoe UI", 16, "bold"), bg=bg, fg=C["text"]).pack(pady=(0, 5))
        Label(frame, text=self.lang.t("about_version"),
              font=("Segoe UI", 10), bg=bg, fg=C["text_secondary"]).pack(pady=(0, 15))
        Frame(frame, height=1, bg=C["border"]).pack(fill="x", pady=(0, 15))

        desc = self.lang.t("about_desc")
        Label(frame, text=desc, font=("Segoe UI", 9),
              bg=bg, fg=C["text"], justify="center").pack(pady=(0, 15))

        dev = Frame(frame, bg=bg)
        dev.pack(pady=(0, 5))
        Label(dev, text=self.lang.t("about_created_by") + " ", font=("Segoe UI", 10),
              bg=bg, fg=C["text"]).pack(side="left")
        Label(dev, text="DevOsmar", font=("Segoe UI", 10, "bold"),
              bg=bg, fg=C["primary"]).pack(side="left")

        Label(frame, text=self.lang.t("about_oss"),
              font=("Segoe UI", 9, "italic"), bg=bg, fg=C["text_secondary"]).pack(pady=(10, 15))
        Frame(frame, height=1, bg=C["border"]).pack(fill="x", pady=(0, 15))

        Label(frame, text=self.lang.t("about_donate"),
              font=("Segoe UI", 11, "bold"), bg=bg, fg=C["text"]).pack(pady=(0, 10))

        def _open_url(url):
            import webbrowser
            webbrowser.open(url)

        def _add_hover(w, hov, norm):
            w.bind("<Enter>", lambda e: w.configure(bg=hov))
            w.bind("<Leave>", lambda e: w.configure(bg=norm))

        paypal = Button(frame, text="💸 Doar com PayPal",
                        font=("Segoe UI", 10, "bold"),
                        bg="#00457C", fg="white",
                        activebackground="#003161", activeforeground="white",
                        cursor="hand2", padx=15, pady=6, borderwidth=0,
                        command=lambda: _open_url(PAYPAL_DONATION_URL))
        paypal.pack(pady=(0, 8))
        _add_hover(paypal, "#003161", "#00457C")
        ToolTip(paypal, self.lang.t("tip_paypal"))

        pix_btn = Button(frame, text="🇧🇷 Doar com PIX",
                         font=("Segoe UI", 10, "bold"),
                         bg="#33a854", fg="white",
                         activebackground="#2a8f47", activeforeground="white",
                         cursor="hand2", padx=15, pady=6, borderwidth=0,
                         command=lambda: self._copy_pix_donation())
        pix_btn.pack(pady=(0, 5))
        _add_hover(pix_btn, "#2a8f47", "#33a854")
        ToolTip(pix_btn, self.lang.t("tip_pix"))

        pix_key_f = Frame(frame, bg=bg)
        pix_key_f.pack(pady=(0, 15))
        Label(pix_key_f, text=self.lang.t("pix_label"), font=("Segoe UI", 8),
              bg=bg, fg="#888888").pack(side="left")
        Label(pix_key_f, text=PIX_KEY, font=("Consolas", 9, "bold"),
              bg=bg, fg="#2a8f47").pack(side="left")

        close = Button(frame, text=self.lang.t("btn_close"), font=("Segoe UI", 10),
                       bg="#e0e0e0", fg="#333333",
                       activebackground="#cccccc",
                       padx=20, pady=4, borderwidth=0, command=dialog.destroy)
        close.pack(pady=(5, 0))
        _add_hover(close, "#cccccc", "#e0e0e0")
        ToolTip(close, self.lang.t("tip_close"))

    def _copy_pix_donation(self):
        self.root.clipboard_clear()
        self.root.clipboard_append(PIX_KEY)
        title = self.lang.t("pix_copied_title")
        msg = self.lang.t("pix_copied_msg", key=PIX_KEY)
        messagebox.showinfo(title, msg)


# =============================================================================
# Entry Point
# =============================================================================

def main():
    if not HAS_OPENPYXL:
        # Console warning kept in Portuguese as it's a dev message
        print("AVISO: openpyxl não instalado.\n"
              "A exportação para XLSX não estará disponível.\n"
              "Instale com: pip install openpyxl\n")

    root = Tk()
    style = ttk.Style(root)
    for theme in ("clam", "vista", "xpnative", "winnative", "alt"):
        if theme in style.theme_names():
            try:
                style.theme_use(theme)
                break
            except TclError:
                continue

    app = ConverterApp(root)
    ensure_dependencies(parent=root, lang=app.lang)
    root.mainloop()


if __name__ == "__main__":
    main()
